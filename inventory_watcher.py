#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
inventory_watcher.py — مراقب مخزون النور
يقرأ ملفات الإكسل كل دقيقة ويحدّث HTML + JSON + يرفع على GitHub
"""

import os
import sys
import json
import time
import subprocess
import traceback
from datetime import datetime, timedelta
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

INVENTORY_DIR = r'D:\Mostafa Ibrahim\جرد يومي'
INVENTORY_FILE = os.path.join(INVENTORY_DIR, 'تقارير  ارصدة المخازن.xlsx')
INVOICES_DIR = r'D:\Mostafa Ibrahim\آخر فاتورة للعميل_files'
INVOICES_FILE = os.path.join(INVOICES_DIR, 'اخر فاتورة لعميل.xlsx')
REPO_DIR = r'C:\Users\GoldenTech\inventory-nour'
HTML_FILE = os.path.join(REPO_DIR, 'index.html')
JSON_FILE = os.path.join(REPO_DIR, 'inventory_data.json')
LOG_FILE = os.path.join(REPO_DIR, 'watcher.log')

os.makedirs(REPO_DIR, exist_ok=True)

import warnings
warnings.filterwarnings('ignore')

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed", file=sys.stderr)
    sys.exit(1)


def log(msg):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f'[{timestamp}] {msg}'
    print(line)
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(line + '\n')
    except:
        pass


def safe_float(val):
    if val is None or val == '':
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


def read_inventory():
    if not os.path.exists(INVENTORY_FILE):
        return [], ''
    try:
        wb = openpyxl.load_workbook(INVENTORY_FILE, data_only=True)
        ws = wb.active
        items = []
        print_date = ''
        try:
            cell_val = str(ws.cell(row=2, column=1).value or '')
            if 'Print Date:' in cell_val:
                print_date = cell_val.replace('Print Date:', '').strip()
        except:
            pass
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0] and not row[2]:
                continue
            items.append({
                'store': str(row[0] or '').strip(),
                'code': str(row[1] or '').strip(),
                'name': str(row[2] or '').strip(),
                'unit': str(row[3] or '').strip(),
                'balance': safe_float(row[4]),
                'detailed': str(row[5] or '').strip(),
                'detained': safe_float(row[6]),
                'group': str(row[7] or '').strip(),
                'supplier': str(row[8] or '').strip(),
            })
        wb.close()
        items.sort(key=lambda x: (x['group'], x['name']))
        return items, print_date
    except Exception as e:
        log(f'ERROR inventory: {e}')
        return [], ''


def read_invoices():
    """Read invoice data and compute analytics"""
    if not os.path.exists(INVOICES_FILE):
        log(f'WARNING: Invoices file not found: {INVOICES_FILE}')
        return {'items': [], 'top_customers': {}, 'item_stats': {}, 'print_date': ''}

    try:
        wb = openpyxl.load_workbook(INVOICES_FILE, data_only=True, read_only=True)
        ws = wb.active

        print_date = ''
        item_data = defaultdict(lambda: {
            'name': '', 'code': '', 'unit': '',
            'total_out': 0, 'total_in': 0, 'total_revenue': 0,
            'invoice_count': 0, 'prices': [], 'customers': defaultdict(float),
            'daily_out': defaultdict(float), 'daily_customers': defaultdict(lambda: defaultdict(float)), 'transactions': []
        })

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0] or not row[1]:
                continue

            code = str(row[0] or '').strip()
            name = str(row[1] or '').strip()
            unit = str(row[2] or '').strip()
            qty_in = safe_float(row[3])
            qty_out = safe_float(row[4])
            price = safe_float(row[5])
            total = safe_float(row[7])
            bayan = str(row[10] or '').strip()
            date_str = str(row[11] or '').strip()
            invoice_num = str(row[12] or '').strip()

            # Parse date
            try:
                if ' ' in date_str:
                    date_part = date_str.split(' ')[0]
                else:
                    date_part = date_str
                dt = datetime.strptime(date_part, '%Y-%m-%d')
            except:
                dt = None

            item = item_data[code]
            item['name'] = name
            item['code'] = code
            item['unit'] = unit
            item['total_out'] += qty_out
            item['total_in'] += qty_in
            item['total_revenue'] += total
            item['invoice_count'] += 1
            if price > 0:
                item['prices'].append(price)

            # Extract customer name from bayan
            # Format: "فاتورة بيع للعميل <name> رقم <num>"
            customer = ''
            if 'للعميل' in bayan:
                try:
                    after = bayan.split('للعميل')[1]
                    if 'رقم' in after:
                        customer = after.split('رقم')[0].strip()
                    else:
                        customer = after.strip()
                except:
                    customer = bayan
            elif bayan:
                customer = bayan

            if customer:
                item['customers'][customer] += qty_out

            if dt:
                date_key = dt.strftime('%Y-%m-%d')
                item['daily_out'][date_key] += qty_out
                if customer:
                    item['daily_customers'][date_key][customer] += qty_out

        wb.close()

        # Build results
        items_list = []
        for code, data in item_data.items():
            avg_price = 0
            if data['prices']:
                avg_price = sum(data['prices']) / len(data['prices'])

            # Latest price (last transaction)
            latest_price = data['prices'][-1] if data['prices'] else 0

            # Top customers for this item
            top_customers = sorted(data['customers'].items(), key=lambda x: x[1], reverse=True)[:5]

            items_list.append({
                'code': code,
                'name': data['name'],
                'unit': data['unit'],
                'total_out': round(data['total_out'], 3),
                'total_in': round(data['total_in'], 3),
                'total_revenue': round(data['total_revenue'], 2),
                'invoice_count': data['invoice_count'],
                'avg_price': round(avg_price, 3),
                'latest_price': round(latest_price, 3),
                'top_customers': [{'name': c, 'qty': round(q, 3)} for c, q in top_customers],
                'daily_out': dict(data['daily_out']),
                'daily_customers': {d: dict(custs) for d, custs in data['daily_customers'].items()},
            })

        # Sort by total_out descending
        items_list.sort(key=lambda x: x['total_out'], reverse=True)

        # Overall top customers
        all_customers = defaultdict(float)
        for item in items_list:
            for c in item['top_customers']:
                all_customers[c['name']] += c['qty']
        top_overall = sorted(all_customers.items(), key=lambda x: x[1], reverse=True)[:10]

        # Daily totals across all items
        daily_totals = defaultdict(float)
        for item in items_list:
            for date, qty in item['daily_out'].items():
                daily_totals[date] += qty
        daily_sorted = sorted(daily_totals.items())

        return {
            'items': items_list,
            'top_customers': [{'name': c, 'qty': round(q, 3)} for c, q in top_overall],
            'daily_totals': [{'date': d, 'qty': round(q, 3)} for d, q in daily_sorted],
            'print_date': print_date,
            'total_transactions': sum(i['invoice_count'] for i in items_list),
        }

    except Exception as e:
        log(f'ERROR invoices: {e}')
        log(traceback.format_exc())
        return {'items': [], 'top_customers': [], 'daily_totals': [], 'print_date': '', 'total_transactions': 0}


def generate_data():
    log('Reading data...')
    inv_items, inv_print_date = read_inventory()
    invoices = read_invoices()

    total_items = len(inv_items)
    total_balance = sum(i['balance'] for i in inv_items)
    low_stock = sum(1 for i in inv_items if 0 < i['balance'] < 50)
    out_of_stock = sum(1 for i in inv_items if i['balance'] <= 0)

    groups = {}
    for item in inv_items:
        g = item['group'] or 'غير مصنف'
        if g not in groups:
            groups[g] = {'count': 0, 'balance': 0}
        groups[g]['count'] += 1
        groups[g]['balance'] += item['balance']

    data = {
        'generated_at': datetime.now().isoformat(),
        'print_date': inv_print_date,
        'invoices_print_date': invoices.get('print_date', ''),
        'store_name': 'النور لتجارة الأعلاف',
        'inventory': inv_items,
        'invoices': invoices,
        'stats': {
            'total_items': total_items,
            'total_balance': round(total_balance, 3),
            'low_stock': low_stock,
            'out_of_stock': out_of_stock,
            'groups': groups,
            'total_transactions': invoices.get('total_transactions', 0),
        }
    }
    return data


def embed_data_into_html(data):
    """Embed a lightweight version of data into HTML (no daily_customers to keep size small)"""
    if not os.path.exists(HTML_FILE):
        return False
    try:
        # Create lightweight copy without daily_customers
        light = json.loads(json.dumps(data))
        if 'invoices' in light and 'items' in light['invoices']:
            for item in light['invoices']['items']:
                item.pop('daily_customers', None)

        with open(HTML_FILE, 'r', encoding='utf-8') as f:
            html = f.read()

        json_str = json.dumps(light, ensure_ascii=False)
        marker_start = '/*__EMBEDDED_DATA__*/'
        marker_end = '/*__END_EMBEDDED_DATA__*/'
        start_idx = html.find(marker_start)
        end_idx = html.find(marker_end)
        if start_idx != -1 and end_idx != -1:
            new_block = f'{marker_start}\nvar EMBEDDED_DATA = {json_str};\n{marker_end}'
            html = html[:start_idx] + new_block + html[end_idx + len(marker_end):]
            with open(HTML_FILE, 'w', encoding='utf-8') as f:
                f.write(html)
            return True
        return False
    except Exception as e:
        log(f'ERROR embed: {e}')
        return False


def git_push():
    try:
        os.chdir(REPO_DIR)
        subprocess.run(['git', 'add', '-A'], capture_output=True, text=True, encoding='utf-8')
        result = subprocess.run(['git', 'diff', '--cached', '--quiet'], capture_output=True, text=True)
        if result.returncode == 0:
            log('No changes to commit')
            return True
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        subprocess.run(['git', 'commit', '-m', f'Auto-update {timestamp}'], capture_output=True, text=True, encoding='utf-8')
        result = subprocess.run(['git', 'push'], capture_output=True, text=True, encoding='utf-8')
        if result.returncode != 0:
            log(f'git push: {result.stderr}')
            return False
        log('Pushed to GitHub')
        return True
    except Exception as e:
        log(f'ERROR git: {e}')
        return False


def run_once():
    try:
        data = generate_data()
        with open(JSON_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        embed_data_into_html(data)
        git_push()
        log(f'Done. Items: {data["stats"]["total_items"]}, Balance: {data["stats"]["total_balance"]}')
        return True
    except Exception as e:
        log(f'FATAL: {e}')
        log(traceback.format_exc())
        return False


def run_loop(interval_seconds=60):
    log(f'=== Watcher Started ===')
    log(f'Inventory: {INVENTORY_FILE}')
    log(f'Invoices: {INVOICES_FILE}')
    log(f'Repo: {REPO_DIR}')
    while True:
        run_once()
        log(f'Next in {interval_seconds}s...')
        try:
            time.sleep(interval_seconds)
        except KeyboardInterrupt:
            break
    log('=== Stopped ===')


if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] == '--once':
        sys.exit(0 if run_once() else 1)
    else:
        run_loop(60)
